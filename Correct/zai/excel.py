"""
excel.py
========
Gestión del DataSet de 3 hojas:
  - transcripciones : audio | zapoteco
  - traducciones    : audio | español
  - palabras_nuevas : zapoteco | español | fuente
"""

from __future__ import annotations
import shutil
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import Workbook

from zai.config import HOJA_ZAPOTECO, HOJA_ESPANOL, HOJA_PALABRAS


# ── Estructuras de datos ─────────────────────────────────────────────

@dataclass
class FilaTranscripcion:
    indice: int                          # 0-based
    audio:  str
    zapoteco: str
    zapoteco_corregido: Optional[str] = None

    @property
    def texto_vigente(self) -> str:
        return self.zapoteco_corregido or self.zapoteco

    @property
    def fue_modificada(self) -> bool:
        return (self.zapoteco_corregido is not None
                and self.zapoteco_corregido != self.zapoteco)


@dataclass
class PalabraNueva:
    zapoteco: str
    espanol:  str
    fuente:   str = ""   # p.ej. "audio_5.wav"


# ── Gestor principal ─────────────────────────────────────────────────

class GestorDataset:
    """
    Carga y guarda el DataSet con 3 hojas.
    Crea las hojas faltantes automáticamente si el archivo
    tiene la estructura antigua (una sola hoja).
    """

    def __init__(self, ruta: str | Path):
        self.ruta = Path(ruta)
        self.filas:    list[FilaTranscripcion] = []
        self.español:  dict[str, str] = {}       # audio → traducción española
        self.palabras: list[PalabraNueva]  = []
        self._cargar()

    # ── Carga ────────────────────────────────────────────────────────

    def _cargar(self) -> None:
        if not self.ruta.exists():
            raise FileNotFoundError(f"Dataset no encontrado: {self.ruta}")

        wb = openpyxl.load_workbook(str(self.ruta))

        # Migración: si el archivo tiene la hoja vieja "Hoja 1", renombrarla
        if HOJA_ZAPOTECO not in wb.sheetnames:
            # Buscar primera hoja con datos y renombrarla
            for nombre in wb.sheetnames:
                ws = wb[nombre]
                if ws.max_row > 1:
                    ws.title = HOJA_ZAPOTECO
                    break

        # Crear hojas faltantes
        for nombre in [HOJA_ZAPOTECO, HOJA_ESPANOL, HOJA_PALABRAS]:
            if nombre not in wb.sheetnames:
                wb.create_sheet(nombre)

        wb.save(str(self.ruta))
        wb.close()

        # Reabrir y leer
        wb = openpyxl.load_workbook(str(self.ruta))

        # Sheet 1: transcripciones
        ws_zap = wb[HOJA_ZAPOTECO]
        self.filas = []
        for idx, fila in enumerate(ws_zap.iter_rows(min_row=1, values_only=True)):
            if not fila:
                continue
            audio = str(fila[0]).strip() if fila[0] else f"fila_{idx + 1}"
            texto = str(fila[1]).strip() if len(fila) > 1 and fila[1] else ""
            if not texto:
                continue
            self.filas.append(FilaTranscripcion(
                indice=idx,
                audio=audio,
                zapoteco=texto,
            ))

        # Sheet 2: traducciones (audio → español)
        ws_esp = wb[HOJA_ESPANOL]
        self.español = {}
        for fila in ws_esp.iter_rows(min_row=1, values_only=True):
            if fila[0] and fila[1]:
                self.español[str(fila[0]).strip()] = str(fila[1]).strip()

        # Sheet 3: palabras nuevas
        ws_pal = wb[HOJA_PALABRAS]
        self.palabras = []
        for fila in ws_pal.iter_rows(min_row=1, values_only=True):
            if fila[0] and fila[1]:
                self.palabras.append(PalabraNueva(
                    zapoteco=str(fila[0]).strip(),
                    espanol=str(fila[1]).strip(),
                    fuente=str(fila[2]).strip() if len(fila) > 2 and fila[2] else "",
                ))

        wb.close()

    # ── Accesores ────────────────────────────────────────────────────

    def obtener_fila(self, n: int) -> FilaTranscripcion:
        return self.filas[n]

    def traduccion_de(self, audio: str) -> str:
        """Traducción española para un audio dado (vacío si no existe)."""
        return self.español.get(audio, "")

    def fila_con_espanol(self, n: int) -> tuple[FilaTranscripcion, str]:
        fila = self.filas[n]
        return fila, self.traduccion_de(fila.audio)

    @property
    def total(self) -> int:
        return len(self.filas)

    # ── Escritura ────────────────────────────────────────────────────

    def guardar_correcciones(
        self,
        filas: list[FilaTranscripcion],
        backup: bool = True,
    ) -> int:
        modificadas = [f for f in filas if f.fue_modificada]
        if not modificadas:
            return 0

        if backup:
            bak = self.ruta.with_suffix(".bak.xlsx")
            if not bak.exists():
                shutil.copy2(self.ruta, bak)

        wb = openpyxl.load_workbook(str(self.ruta))
        ws = wb[HOJA_ZAPOTECO]

        for fila in modificadas:
            row_num = fila.indice + 1          # openpyxl es 1-based
            ws.cell(row=row_num, column=2).value = fila.zapoteco_corregido

        wb.save(str(self.ruta))
        wb.close()
        return len(modificadas)

    def guardar_traduccion(self, audio: str, espanol: str) -> None:
        """Añade o actualiza la traducción española de un audio."""
        self.español[audio] = espanol
        wb = openpyxl.load_workbook(str(self.ruta))
        ws = wb[HOJA_ESPANOL]

        # Buscar fila existente o añadir al final
        encontrada = False
        for row in ws.iter_rows(min_row=1):
            if row[0].value and str(row[0].value).strip() == audio:
                row[1].value = espanol
                encontrada = True
                break
        if not encontrada:
            ws.append([audio, espanol])

        wb.save(str(self.ruta))
        wb.close()

    def guardar_palabra_nueva(self, palabra: PalabraNueva) -> None:
        """Añade una palabra nueva a la hoja de palabras_nuevas."""
        # Evitar duplicados
        existentes = {p.zapoteco.lower() for p in self.palabras}
        if palabra.zapoteco.lower() in existentes:
            return
        self.palabras.append(palabra)

        wb = openpyxl.load_workbook(str(self.ruta))
        ws = wb[HOJA_PALABRAS]
        ws.append([palabra.zapoteco, palabra.espanol, palabra.fuente])
        wb.save(str(self.ruta))
        wb.close()

    def guardar_batch_palabras(self, palabras: list[PalabraNueva]) -> int:
        """Guarda múltiples palabras nuevas evitando duplicados. Retorna cuántas se guardaron."""
        existentes = {p.zapoteco.lower() for p in self.palabras}
        nuevas = [p for p in palabras if p.zapoteco.lower() not in existentes]
        if not nuevas:
            return 0

        wb = openpyxl.load_workbook(str(self.ruta))
        ws = wb[HOJA_PALABRAS]
        for p in nuevas:
            ws.append([p.zapoteco, p.espanol, p.fuente])
            self.palabras.append(p)
            existentes.add(p.zapoteco.lower())

        wb.save(str(self.ruta))
        wb.close()
        return len(nuevas)
