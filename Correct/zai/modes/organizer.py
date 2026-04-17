"""
modes/organizer.py
==================
Organiza palabras nuevas y el diccionario siguiendo el orden
ortográfico del Zapoteco del Istmo (dígrafos como unidades).
"""

from __future__ import annotations
from pathlib import Path
from typing import Optional

import openpyxl

from zai.excel import GestorDataset, PalabraNueva
from zai.context import ContextoLinguistico


# ── Orden zapoteco ───────────────────────────────────────────────────

_ORDEN = [
    "a", "b", "c", "ch", "d", "dx", "e", "f", "g", "gu", "gü",
    "h", "hui", "i", "j", "k", "l", "ll", "m", "mb", "n",
    "nc", "nd", "ng", "nn", "ñ", "o", "p", "q", "qu", "r",
    "rr", "s", "t", "tx", "u", "v", "x", "xh", "xp", "y", "z", "zh",
]

# Reemplazar dígrafos por un carácter único para comparación
_DIGRAFO_MAP: dict[str, str] = {}
_UNICODE_BASE = 0xE000   # área de uso privado

for _i, _d in enumerate(_ORDEN):
    _DIGRAFO_MAP[_d] = chr(_UNICODE_BASE + _i)


def _clave_orden(palabra: str) -> str:
    """Genera clave de ordenamiento respetando el alfabeto zapoteco."""
    p = palabra.lower().strip()
    resultado = []
    i = 0
    while i < len(p):
        # Intentar dígrafos de 3 letras primero
        for tam in (3, 2, 1):
            fragmento = p[i:i + tam]
            if fragmento in _DIGRAFO_MAP:
                resultado.append(_DIGRAFO_MAP[fragmento])
                i += tam
                break
        else:
            resultado.append(p[i])
            i += 1
    return "".join(resultado)


# ── Modo Organizador ─────────────────────────────────────────────────

class ModoOrganizador:

    def __init__(self, dataset: GestorDataset, contexto: ContextoLinguistico):
        self.ds  = dataset
        self.ctx = contexto

    # ── Ordenamiento ─────────────────────────────────────────────────

    def ordenar(self, palabras: list[PalabraNueva]) -> list[PalabraNueva]:
        return sorted(palabras, key=lambda p: _clave_orden(p.zapoteco))

    def ordenar_diccionario(self) -> list[tuple[str, str]]:
        """Ordena todas las entradas del diccionario cargado."""
        return sorted(self.ctx.diccionario, key=lambda e: _clave_orden(e[0]))

    # ── Exportar ─────────────────────────────────────────────────────

    def exportar_nueva_hoja(
        self,
        palabras:    list[PalabraNueva],
        ruta_excel:  Path,
        nombre_hoja: str = "Palabras Ordenadas",
    ) -> int:
        """Exporta palabras ordenadas a una nueva hoja del Excel."""
        ordenadas = self.ordenar(palabras)

        wb = openpyxl.load_workbook(str(ruta_excel))
        if nombre_hoja in wb.sheetnames:
            del wb[nombre_hoja]
        ws = wb.create_sheet(nombre_hoja)
        ws.append(["Zapoteco", "Español", "Fuente"])
        for p in ordenadas:
            ws.append([p.zapoteco, p.espanol, p.fuente])
        wb.save(str(ruta_excel))
        wb.close()
        return len(ordenadas)

    def exportar_excel_independiente(
        self,
        palabras: list[PalabraNueva],
        ruta:     Path,
    ) -> int:
        """Crea un archivo Excel independiente con las palabras ordenadas."""
        ordenadas = self.ordenar(palabras)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Palabras Nuevas"
        ws.append(["Zapoteco", "Español", "Fuente"])
        for p in ordenadas:
            ws.append([p.zapoteco, p.espanol, p.fuente])
        wb.save(str(ruta))
        return len(ordenadas)

    def integrar_al_diccionario(
        self,
        palabras_nuevas: list[PalabraNueva],
        ruta_diccionario: Path,
    ) -> int:
        """
        Integra palabras nuevas al diccionario, reordena todo
        y escribe de vuelta al archivo.
        """
        # Combinar existentes + nuevas
        existentes = {e[0].lower() for e in self.ctx.diccionario}
        agregar = [p for p in palabras_nuevas if p.zapoteco.lower() not in existentes]

        todo = list(self.ctx.diccionario) + [(p.zapoteco, p.espanol) for p in agregar]
        ordenado = sorted(todo, key=lambda e: _clave_orden(e[0]))

        wb = openpyxl.load_workbook(str(ruta_diccionario))
        ws = wb.active
        # Limpiar y reescribir
        ws.delete_rows(1, ws.max_row)
        ws.append(["ZAPOTECO", "ESPAÑOL"])
        for zap, esp in ordenado:
            ws.append([zap, esp])
        wb.save(str(ruta_diccionario))
        wb.close()
        return len(agregar)
